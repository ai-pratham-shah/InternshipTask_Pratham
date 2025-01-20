import openpyxl

# Load or create the Excel workbook
def load_workbook():
    try:
        wb = openpyxl.load_workbook("data.xlsx")
        return wb
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        wb = openpyxl.Workbook()
        wb.create_sheet("Records")
        wb.save("data.xlsx")
        return wb

# Save the data to Excel file
def save_to_excel(records):
    try:
        wb = load_workbook()
        ws = wb["Records"]

        # Clear existing data before writing new data
        ws.delete_rows(1, ws.max_row)  # Clear all data

        # Write headers
        ws.append(["Country", "State", "City"])

        # Write records in the desired format
        for country, states in records.items():
            for state, cities in states.items():
                first_row = True
                for city in cities:
                    if first_row:
                        ws.append([country, state, city])
                        first_row = False
                    else:
                        ws.append(["", "", city])  

        # Save the file
        wb.save("data.xlsx")
        print("Data saved to Excel successfully.")
    except Exception as e:
        print(f"Error occurred while saving data to Excel: {e}")

# Modify add_data function to handle multiple cities in the same row
def add_data():
    try:
        country = input("Enter country: ").strip().title().replace(" ","")
        if not country.isalpha():
            print(" Invalid Country ")
            return add_data()

        state = input("Enter state: ").strip().title().replace(" ","")
        if not state.isalpha():
            print(" Invalid state ")
            return add_data()

        city = input("Enter city: ").strip().title().replace(" ","")
        if not city.isalpha():
            print(" Invalid city ")
            return add_data()

        if country not in records:
            records[country] = {}
        
        if state not in records[country]:
            records[country][state] = []
        
        
        if city not in records[country][state]:
            records[country][state].append(city)
            print(f"Record added: Country = {country}, State = {state}, City = {city}")
            save_to_excel(records)  # Save to Excel after adding the city
        else:
            print(f"City {city} already exists under Country = {country}, State = {state}.")

        choice = input("Do you want to add another record? (yes/no): ").strip().lower()
        if choice == "yes":
            add_data()
        elif choice == "no":
            print("Returning to the main menu.")
        else:
            print("Invalid input. Returning to the main menu.")

    except Exception as e:
        print(f"Error occurred while adding record: {e}")

# Show all data
def show_data():
    try:
        if not records:
            print("No records found.")
        else:
            print("All Records:")
            for country, states in records.items():
                print(f"Country: {country}")
                for state, cities in states.items():
                    print(f"  State: {state}, Cities: {', '.join(cities)}")
    except Exception as e:
        print(f"Error occurred while showing records: {e}")

# Update data
def update_data():
    try:
        print("What would you like to update (Country, State, City)?")
        choice = input("Enter your choice (Country, State, City): ").strip().title()

        if choice == "Country":
            old_country = input("Enter the current country name: ").strip().title()
            if old_country in records:
                new_country = input(f"Enter the new name for the country {old_country}: ").strip().title()
                records[new_country] = records.pop(old_country)
                print(f"Country {old_country} updated to {new_country}.")
                save_to_excel(records)  # Save to Excel here
            else:
                print(f"Country {old_country} not found.")

        elif choice == "State":
            country = input("Enter the country for the state: ").strip().title()
            if country in records:
                old_state = input("Enter the current state name: ").strip().title()
                if old_state in records[country]:
                    new_state = input(f"Enter the new name for the state {old_state}: ").strip().title()
                    records[country][new_state] = records[country].pop(old_state)
                    print(f"State {old_state} updated to {new_state} under Country = {country}.")
                    save_to_excel(records)  # Save to Excel here
                else:
                    print(f"State {old_state} not found under Country = {country}.")
            else:
                print(f"Country {country} not found.")

        elif choice == "City":
            country = input("Enter the country for the city: ").strip().title()
            if country in records:
                state = input("Enter the state for the city: ").strip().title()
                if state in records[country]:
                    old_city = input("Enter the current city name: ").strip().title()
                    if old_city in records[country][state]:
                        new_city = input(f"Enter the new name for the city {old_city}: ").strip().title()
                        records[country][state].remove(old_city)
                        records[country][state].append(new_city)
                        print(f"City {old_city} updated to {new_city} under Country = {country}, State = {state}.")
                        save_to_excel(records)  # Save to Excel here
                    else:
                        print(f"City {old_city} not found under Country = {country}, State = {state}.")
                else:
                    print(f"State {state} not found under Country = {country}.")
            else:
                print(f"Country {country} not found.")

        else:
            print("Invalid choice. Please select Country, State, or City.")
    except Exception as e:
        print(f"Error occurred while updating record: {e}")

# Delete data
def delete_data():
    try:
        print("What would you like to delete? (Country, State, City)")
        choice = input("Enter your choice (Country, State, City): ").strip().title()

        if choice == "Country":
            country = input("Enter the country to delete: ").strip().title()
            if country in records:
                del records[country]
                print(f"All records for Country = {country} deleted successfully.")
                save_to_excel(records)  # Save to Excel here
            else:
                print(f"Country {country} not found.")

        elif choice == "State":
            country = input("Enter the country for the state: ").strip().title()
            if country in records:
                state = input("Enter the state to delete: ").strip().title()
                if state in records[country]:
                    del records[country][state]
                    print(f"State {state} deleted successfully under Country = {country}.")
                    save_to_excel(records)  # Save to Excel here
                else:
                    print(f"State {state} not found under Country = {country}.")
            else:
                print(f"Country {country} not found.")

        elif choice == "City":
            country = input("Enter the country for the city: ").strip().title()
            if country in records:
                state = input("Enter the state for the city: ").strip().title()
                if state in records[country]:
                    city = input("Enter the city to delete: ").strip().title()
                    if city in records[country][state]:
                        records[country][state].remove(city)
                        print(f"City {city} deleted successfully under Country = {country}, State = {state}.")
                        save_to_excel(records)  # Save to Excel here
                    else:
                        print(f"City {city} not found under Country = {country}, State = {state}.")
                else:
                    print(f"State {state} not found under Country = {country}.")
            else:
                print(f"Country {country} not found.")

        else:
            print("Invalid choice. Please choose either Country, State, or City.")
    except Exception as e:
        print(f"Error occurred while deleting record: {e}")

# Main menu function
def main_menu():
    try:
        print("\nMenu:")
        print("1. Enter Country, State, City")
        print("2. Show All Records")
        print("3. Update Record")
        print("4. Delete Record")
        print("5. Exit")

        choice = input("Enter your choice: ").strip()

        if choice == "1":
            add_data()
            main_menu()
        elif choice == "2":
            show_data()
            main_menu()
        elif choice == "3":
            update_data()
            main_menu()
        elif choice == "4":
            delete_data()
            main_menu()
        elif choice == "5":
            print("Exiting the program. Goodbye!")
        else:
            print("Invalid choice. Please try again.")
            main_menu()
    except Exception as e:
        print(f"Error occurred: {e}")

# Initialize the records dictionary
records = {}
main_menu()

